# rubicon_bot.py — Rubicon Production: RU/UZ/EN + Excel + webhook (Render)
# Требует: python-telegram-bot[webhooks]==20.7, openpyxl
# Новое: REQ-ID, статусы, /stats, /list, ADMIN_CHAT_ID, валидация телефона/почты, защита /admin

import os
import re
import logging
from datetime import datetime, timedelta
from typing import Dict, Any, Optional, List, Tuple

from telegram import (
    Update, InlineKeyboardMarkup, InlineKeyboardButton, User
)
from telegram.ext import (
    ApplicationBuilder, CommandHandler, CallbackQueryHandler,
    MessageHandler, ContextTypes, filters
)

# ── ENV / CONFIG ──────────────────────────────────────────────────────────────
BOT_TOKEN = os.getenv("BOT_TOKEN")
if not BOT_TOKEN:
    raise RuntimeError("Переменная окружения BOT_TOKEN не задана.")

# Админ — стабильно через переменные окружения
ADMIN_ID_ENV = os.getenv("ADMIN_ID")  # chat_id (число)
ADMIN_USERNAME_ENV = (os.getenv("ADMIN_USERNAME") or "").strip().lstrip("@").lower()
ADMIN_KEY_ENV = os.getenv("ADMIN_KEY", "").strip()  # optional секрет для /admin <key>

# Опционально: командный чат/канал (обычно отрицательное число для групп)
# Пример: -1001234567890
ADMIN_CHAT_ID_ENV = os.getenv("ADMIN_CHAT_ID")

ADMIN_FILE = "admin_id.txt"    # кэш (на free Render может пропадать)
EXCEL_FILE = "requests.xlsx"   # локальный Excel (/export, /list, /stats)

ADMIN_ID: Optional[int] = None
user_lang: Dict[int, str] = {}
forms: Dict[int, Dict[str, Any]] = {}  # user_id -> {"step": int, "data": {...}}
REQ_PREFIX = "REQ-"

logging.basicConfig(
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
    level=logging.INFO,
)
log = logging.getLogger("rubicon")

# ── ТЕКСТЫ ─────────────────────────────────────────────────────────────────────
T = {
    "ru": {
        "welcome": "👋 <b>Rubicon Production</b>\nВыберите язык и нажмите «Заполнить заявку».",
        "choose_lang": "🌐 Выберите язык:",
        "btn_ru": "Русский 🇷🇺", "btn_uz": "Oʻzbekcha 🇺🇿", "btn_en": "English 🇬🇧",
        "lang_set": "✅ Язык переключён на русский. Нажмите «📝 Заполнить заявку».",
        "btn_form": "📝 Заполнить заявку", "btn_lang": "🌐 Сменить язык",

        "form_started": "📝 Начинаем заявку.\n<b>Шаг 1/6</b>\nВведите <b>ФИО</b> и <b>название компании</b> (одной строкой).",
        "ask_phone":  "📞 <b>Шаг 2/6</b>\nВведите номер телефона (с кодом страны).",
        "ask_tg":     "📨 <b>Шаг 3/6</b>\nВаш Telegram-ник (с @ или без).",
        "ask_task":   "🎯 <b>Шаг 4/6</b>\nКратко опишите задачу/проект.",
        "ask_email":  "✉️ <b>Шаг 5/6</b>\nВаш e-mail.",
        "ask_confirm":"🔎 <b>Шаг 6/6</b>\nПроверьте данные и нажмите «Подтвердить»:",

        "btn_confirm": "✅ Подтвердить и отправить", "btn_cancel": "✖️ Отмена",
        "sent_user": "✅ Заявка отправлена! Мы свяжемся с вами.",
        "sent_admin": "📩 Новая заявка:",

        "not_admin": "Админ не назначен. Используйте /whoami и занесите ADMIN_ID в Render.",
        "admin_set": "✅ Вы назначены администратором и будете получать заявки.",
        "cancelled": "❌ Отменено. Для новой заявки нажмите «📝 Заполнить заявку».",
        "export_ok": "📎 Отправляю текущий Excel с заявками.",
        "export_none": "🗂 Файл ещё не создан (нет заявок).",
        "cleared": "🧹 Готово: локальный Excel удалён.",
        "no_rights": "Только администратор может использовать эту команду.",
        "whoami": "🆔 Ваш chat_id: <code>{}</code>\nДобавьте его в Render → Environment как <code>ADMIN_ID</code>.",

        "bad_phone": "⚠️ Похоже, телефон в странном формате. Я приведу к виду <code>+999…</code>.",
        "bad_email": "⚠️ Похоже, e-mail указан неверно. Проверьте, пожалуйста.",

        "status_set": "Статус заявки {} → <b>{}</b>",
        "stats": "📊 Статистика:\nСегодня: <b>{}</b>\n7 дней: <b>{}</b>\nВсего: <b>{}</b>",
        "list_header": "🗂 Последние заявки:",
    },
    "uz": {
        "welcome": "👋 <b>Rubicon Production</b>\nTilni tanlang va «Ariza yuborish» tugmasini bosing.",
        "choose_lang": "🌐 Tilni tanlang:",
        "btn_ru": "Ruscha 🇷🇺", "btn_uz": "Oʻzbekcha 🇺🇿", "btn_en": "Inglizcha 🇬🇧",
        "lang_set": "✅ Til oʻzbekchaga oʻzgartirildi. «📝 Ariza yuborish» ni bosing.",
        "btn_form": "📝 Ariza yuborish", "btn_lang": "🌐 Tilni almashtirish",

        "form_started": "📝 Ariza boshlandi.\n<b>1/6</b>\nF.I.Sh. va kompaniya nomi (bir qatorda).",
        "ask_phone":  "📞 <b>2/6</b>\nTelefon raqamingiz (mamlakat kodi bilan).",
        "ask_tg":     "📨 <b>3/6</b>\nTelegram nick (@ bilan yoki bo‘lmasligi mumkin).",
        "ask_task":   "🎯 <b>4/6</b>\nVazifa/proyektni qisqa yozing.",
        "ask_email":  "✉️ <b>5/6</b>\nE-mail.",
        "ask_confirm":"🔎 <b>6/6</b>\nTekshirib «Tasdiqlash» tugmasini bosing:",

        "btn_confirm": "✅ Tasdiqlash", "btn_cancel": "✖️ Bekor qilish",
        "sent_user": "✅ Arizangiz yuborildi!",
        "sent_admin": "📩 Yangi ariza:",

        "not_admin": "Admin belgilanmagan. /whoami yuboring va ADMIN_ID ni qo‘shing.",
        "admin_set": "✅ Admin sifatida belgilandingiz.",
        "cancelled": "❌ Bekor qilindi.",
        "export_ok": "📎 Excel faylini yuboraman.",
        "export_none": "🗂 Fayl hali yaratilmagan.",
        "cleared": "🧹 Excel tozalandi.",
        "no_rights": "Bu buyruqni faqat admin ishlatishi mumkin.",
        "whoami": "🆔 Sizning chat_id: <code>{}</code>",

        "bad_phone": "⚠️ Telefon formati g‘alati. Men uni <code>+999…</code> ko‘rinishga keltiraman.",
        "bad_email": "⚠️ E-mail xato bo‘lishi mumkin. Iltimos, tekshiring.",

        "status_set": "Ariza holati {} → <b>{}</b>",
        "stats": "📊 Statistika:\nBugun: <b>{}</b>\n7 kun: <b>{}</b>\nJami: <b>{}</b>",
        "list_header": "🗂 So‘nggi arizalar:",
    },
    "en": {
        "welcome": "👋 <b>Rubicon Production</b>\nChoose language and tap “Submit request”.",
        "choose_lang": "🌐 Choose language:",
        "btn_ru": "Russian 🇷🇺", "btn_uz": "Uzbek 🇺🇿", "btn_en": "English 🇬🇧",
        "lang_set": "✅ Language set to English. Tap “📝 Submit request”.",
        "btn_form": "📝 Submit request", "btn_lang": "🌐 Change language",

        "form_started": "📝 Let’s start.\n<b>Step 1/6</b>\nFull name + company (one line).",
        "ask_phone":  "📞 <b>Step 2/6</b>\nPhone number (with country code).",
        "ask_tg":     "📨 <b>Step 3/6</b>\nTelegram handle (with or without @).",
        "ask_task":   "🎯 <b>Step 4/6</b>\nShort description of your task/project.",
        "ask_email":  "✉️ <b>Step 5/6</b>\nYour e-mail.",
        "ask_confirm":"🔎 <b>Step 6/6</b>\nCheck details and press “Confirm”:",

        "btn_confirm": "✅ Confirm & send", "btn_cancel": "✖️ Cancel",
        "sent_user": "✅ Request sent!",
        "sent_admin": "📩 New request:",

        "not_admin": "Admin not set. Use /whoami and set ADMIN_ID in Render.",
        "admin_set": "✅ You are set as admin.",
        "cancelled": "❌ Cancelled.",
        "export_ok": "📎 Sending Excel.",
        "export_none": "🗂 File not created yet.",
        "cleared": "🧹 Excel removed.",
        "no_rights": "Only the admin can use this command.",
        "whoami": "🆔 Your chat_id: <code>{}</code>",

        "bad_phone": "⚠️ Phone looks odd. I’ll normalize it to <code>+999…</code>.",
        "bad_email": "⚠️ E-mail seems invalid. Please check.",

        "status_set": "Request {} status → <b>{}</b>",
        "stats": "📊 Stats:\nToday: <b>{}</b>\n7 days: <b>{}</b>\nTotal: <b>{}</b>",
        "list_header": "🗂 Latest requests:",
    }
}

STATUSES = {
    "INPROG": "✅ В работе",
    "WAIT":   "⏳ Ожидание",
    "REJ":    "❌ Отказ",
}

# ── HELPERS ───────────────────────────────────────────────────────────────────
def get_lang(uid: int) -> str:
    return user_lang.get(uid, "ru")

def main_menu(lang: str) -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup([[
        InlineKeyboardButton(T[lang]["btn_form"], callback_data="form:start"),
    ], [
        InlineKeyboardButton(T[lang]["btn_lang"], callback_data="lang:open"),
    ]])

def lang_menu(lang: str) -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup([[
        InlineKeyboardButton(T[lang]["btn_ru"], callback_data="set_lang:ru"),
        InlineKeyboardButton(T[lang]["btn_uz"], callback_data="set_lang:uz"),
        InlineKeyboardButton(T[lang]["btn_en"], callback_data="set_lang:en"),
    ]])

def render_card(req_id: str, d: Dict[str, str]) -> str:
    return (
        (f"<b>ID:</b> {req_id}\n" if req_id else "") +
        f"<b>ФИО+компания:</b> {d.get('fio_company','—')}\n"
        f"<b>Телефон:</b> {d.get('phone','—')}\n"
        f"<b>Telegram:</b> @{d.get('tg','—').lstrip('@')}\n"
        f"<b>Задача:</b> {d.get('task','—')}\n"
        f"<b>Email:</b> {d.get('email','—')}"
    )

def load_admin_id_from_file() -> Optional[int]:
    if os.path.exists(ADMIN_FILE):
        try:
            return int(open(ADMIN_FILE, "r", encoding="utf-8").read().strip())
        except Exception:
            return None
    return None

def save_admin_id_to_file(admin_id: int):
    try:
        with open(ADMIN_FILE, "w", encoding="utf-8") as f:
            f.write(str(admin_id))
    except Exception as e:
        log.warning("Не удалось сохранить admin_id.txt: %s", e)

def bootstrap_admin_from_env():
    global ADMIN_ID
    if ADMIN_ID_ENV:
        try:
            ADMIN_ID = int(ADMIN_ID_ENV)
            log.info("ADMIN_ID из ENV: %s", ADMIN_ID)
            return
        except ValueError:
            log.warning("ADMIN_ID в ENV не число: %r", ADMIN_ID_ENV)
    file_id = load_admin_id_from_file()
    if file_id:
        ADMIN_ID = file_id
        log.info("ADMIN_ID из файла: %s", ADMIN_ID)

def maybe_auto_claim_admin(user: User):
    global ADMIN_ID
    if ADMIN_ID is None and ADMIN_USERNAME_ENV:
        if (user.username or "").lower() == ADMIN_USERNAME_ENV:
            ADMIN_ID = user.id
            save_admin_id_to_file(ADMIN_ID)
            log.info("ADMIN_ID назначен по username=%s: %s", ADMIN_USERNAME_ENV, ADMIN_ID)
            return True
    return False

def get_admin_chat_id() -> Optional[int]:
    if not ADMIN_CHAT_ID_ENV:
        return None
    try:
        return int(ADMIN_CHAT_ID_ENV)
    except ValueError:
        return None

# ── Excel ─────────────────────────────────────────────────────────────────────
def _excel_headers() -> List[str]:
    return [
        "ReqID", "TimestampUTC", "Lang", "FIO+Company", "Phone",
        "Telegram", "Task", "Email", "UserID", "Username",
        "Status", "StatusBy", "StatusAtUTC"
    ]

def excel_append_new(lang: str, req_id: str, data: Dict[str, str], user: User):
    from openpyxl import Workbook, load_workbook
    from openpyxl.utils import get_column_letter

    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Requests"
        ws.append(_excel_headers())
        wb.save(EXCEL_FILE)

    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    row = [
        req_id,
        datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S"),
        lang.upper(),
        data.get("fio_company", ""),
        data.get("phone", ""),
        f"@{data.get('tg','').lstrip('@')}",
        data.get("task", ""),
        data.get("email", ""),
        user.id,
        f"@{user.username}" if user.username else "",
        "", "", ""  # Status, StatusBy, StatusAtUTC (пока пусто)
    ]
    ws.append(row)

    if ws.max_row == 2:  # первая запись → ширины
        for col in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col)
            max_len = max(len(str(cell.value)) if cell.value else 0 for cell in ws[col_letter])
            ws.column_dimensions[col_letter].width = min(max(12, max_len + 2), 60)

    wb.save(EXCEL_FILE)

def excel_set_status(req_id: str, status_code: str, by_user: User):
    from openpyxl import load_workbook
    if not os.path.exists(EXCEL_FILE):
        return False
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    # ищем строку по ReqID
    for r in range(2, ws.max_row + 1):
        if (ws.cell(row=r, column=1).value or "").strip() == req_id:
            ws.cell(row=r, column=11).value = STATUSES.get(status_code, status_code)  # Status
            ws.cell(row=r, column=12).value = f"@{by_user.username}" if by_user.username else by_user.id  # StatusBy
            ws.cell(row=r, column=13).value = datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S")
            wb.save(EXCEL_FILE)
            return True
    return False

def excel_count(period: str) -> int:
    """
    period: 'today' | '7d' | 'all'
    """
    from openpyxl import load_workbook
    if not os.path.exists(EXCEL_FILE):
        return 0
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    if period == "all":
        return max(ws.max_row - 1, 0)

    now = datetime.utcnow()
    cnt = 0
    for r in range(2, ws.max_row + 1):
        ts = ws.cell(row=r, column=2).value  # TimestampUTC
        if not ts: 
            continue
        try:
            dt = datetime.strptime(str(ts), "%Y-%m-%d %H:%M:%S")
        except Exception:
            continue
        if period == "today":
            if dt.date() == now.date(): 
                cnt += 1
        elif period == "7d":
            if now - dt <= timedelta(days=7): 
                cnt += 1
    return cnt

def excel_last(n: int = 5) -> List[Tuple[str, str, str]]:
    """
    Возвращает последние n заявок: [(ReqID, TimestampUTC, FIO+Company), ...]
    """
    from openpyxl import load_workbook
    if not os.path.exists(EXCEL_FILE):
        return []
    wb = load_workbook(EXCEL_FILE)
    ws = wb.active
    rows = []
    for r in range(2, ws.max_row + 1):
        req_id = ws.cell(row=r, column=1).value or ""
        ts = ws.cell(row=r, column=2).value or ""
        fio = ws.cell(row=r, column=4).value or ""
        rows.append((str(req_id), str(ts), str(fio)))
    return rows[-n:]

def next_request_id() -> str:
    """
    Считает следующий ID на основе количества строк в Excel.
    Если файла нет — REQ-0001.
    """
    try:
        from openpyxl import load_workbook
        if not os.path.exists(EXCEL_FILE):
            return f"{REQ_PREFIX}0001"
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        num = max(ws.max_row - 1, 0) + 1
        return f"{REQ_PREFIX}{num:04d}"
    except Exception:
        # fallback на время
        return f"{REQ_PREFIX}{datetime.utcnow().strftime('%m%d%H%M%S')}"
# ── Валидация ────────────────────────────────────────────────────────────────
def normalize_phone(raw: str) -> Tuple[str, bool]:
    """
    Оставляем только цифры и ведущий '+'. Если '+' нет — добавим.
    Возвращаем (нормализованный, был_подозрительный).
    """
    s = raw.strip()
    was_weird = False
    if not s:
        return s, False
    digits = re.sub(r"\D", "", s)
    if not digits:
        return raw, True
    if s.startswith("+"):
        return f"+{digits}", False
    # если начиналось без '+', нормализуем
    was_weird = True
    return f"+{digits}", was_weird

def email_valid(addr: str) -> bool:
    return bool(re.match(r"^[^@\s]+@[^@\s]+\.[^@\s]+$", addr.strip(), re.IGNORECASE))

# ── Хэндлеры ──────────────────────────────────────────────────────────────────
async def cmd_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user
    maybe_auto_claim_admin(user)
    lang = get_lang(user.id)
    await update.message.reply_text(
        f"{T[lang]['welcome']}\n\n{T[lang]['choose_lang']}",
        parse_mode="HTML",
        reply_markup=lang_menu(lang)
    )

# /admin — защищённая
async def cmd_admin(update: Update, context: ContextTypes.DEFAULT_TYPE):
    global ADMIN_ID
    user = update.effective_user
    lang = get_lang(user.id)

    def username_ok(u: User) -> bool:
        return bool(ADMIN_USERNAME_ENV) and (u.username or "").lower() == ADMIN_USERNAME_ENV
    def key_ok() -> bool:
        return bool(ADMIN_KEY_ENV) and len(context.args) >= 1 and context.args[0] == ADMIN_KEY_ENV

    allowed = False
    if ADMIN_ID is not None:
        allowed = (user.id == ADMIN_ID)
    else:
        allowed = username_ok(user) or key_ok()

    if not allowed:
        await update.message.reply_text(T[lang]["no_rights"]); return

    ADMIN_ID = user.id
    save_admin_id_to_file(ADMIN_ID)
    await update.message.reply_text(T[lang]["admin_set"])

async def cmd_whoami(update: Update, context: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    lang = get_lang(uid)
    await update.message.reply_text(T[lang]["whoami"].format(uid), parse_mode="HTML")

async def cmd_export(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user
    lang = get_lang(user.id)
    if user.id != ADMIN_ID:
        await update.message.reply_text(T[lang]["no_rights"]); return
    if not os.path.exists(EXCEL_FILE):
        await update.message.reply_text(T[lang]["export_none"]); return
    await update.message.reply_text(T[lang]["export_ok"])
    try:
        with open(EXCEL_FILE, "rb") as f:
            await context.bot.send_document(chat_id=update.effective_chat.id, document=f, filename=EXCEL_FILE)
    except Exception as e:
        log.error("Export send failed: %s", e)

async def cmd_clear(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user
    lang = get_lang(user.id)
    if user.id != ADMIN_ID:
        await update.message.reply_text(T[lang]["no_rights"]); return
    try:
        if os.path.exists(EXCEL_FILE):
            os.remove(EXCEL_FILE)
        await update.message.reply_text(T[lang]["cleared"])
    except Exception as e:
        log.error("Clear failed: %s", e)

async def cmd_stats(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user
    lang = get_lang(user.id)
    if user.id != ADMIN_ID:
        await update.message.reply_text(T[lang]["no_rights"]); return
    today = excel_count("today")
    week = excel_count("7d")
    total = excel_count("all")
    await update.message.reply_text(T[lang]["stats"].format(today, week, total), parse_mode="HTML")

async def cmd_list(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user
    lang = get_lang(user.id)
    if user.id != ADMIN_ID:
        await update.message.reply_text(T[lang]["no_rights"]); return
    rows = excel_last(5)
    if not rows:
        await update.message.reply_text("Пока заявок нет."); return
    txt = [T[lang]["list_header"]]
    kb_rows = []
    for req_id, ts, fio in rows:
        txt.append(f"• <b>{req_id}</b> — {fio}  ({ts})")
        # для каждой заявки — быстрые кнопки статуса
        kb_rows.append([
            InlineKeyboardButton("✅ В работе", callback_data=f"st:{req_id}:INPROG"),
            InlineKeyboardButton("⏳ Ожидание", callback_data=f"st:{req_id}:WAIT"),
            InlineKeyboardButton("❌ Отказ",   callback_data=f"st:{req_id}:REJ"),
        ])
    await update.message.reply_text(
        "\n".join(txt), parse_mode="HTML",
        reply_markup=InlineKeyboardMarkup(kb_rows)
    )

async def on_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    user = q.from_user
    maybe_auto_claim_admin(user)
    uid = user.id
    lang = get_lang(uid)
    data = q.data or ""

    if data.startswith("set_lang:"):
        new_lang = data.split(":", 1)[1]
        if new_lang in ("ru", "uz", "en"):
            user_lang[uid] = new_lang
            lang = new_lang
        await q.edit_message_text(T[lang]["lang_set"], parse_mode="HTML", reply_markup=main_menu(lang)); return

    if data == "lang:open":
        await q.edit_message_text(T[lang]["choose_lang"], parse_mode="HTML", reply_markup=lang_menu(lang)); return

    if data == "form:start":
        forms[uid] = {"step": 1, "data": {}}
        await q.edit_message_text(T[lang]["form_started"], parse_mode="HTML"); return

    if data.startswith("st:"):
        # формат: st:<REQ-ID>:<CODE>
        if user.id != ADMIN_ID:
            await q.message.reply_text(T[lang]["no_rights"]); return
        parts = data.split(":")
        if len(parts) == 3:
            req_id, code = parts[1], parts[2]
            ok = excel_set_status(req_id, code, user)
            status_text = STATUSES.get(code, code)
            if ok:
                await q.message.reply_text(T[lang]["status_set"].format(req_id, status_text), parse_mode="HTML")
            else:
                await q.message.reply_text(f"Не нашёл заявку {req_id} в Excel.")

async def on_text(update: Update, context: ContextTypes.DEFAULT_TYPE):
    uid = update.effective_user.id
    user = update.effective_user
    maybe_auto_claim_admin(user)
    lang = get_lang(uid)
    txt = (update.message.text or "").strip()

    if uid not in forms:
        return

    st = forms[uid]["step"]
    d = forms[uid]["data"]

    if st == 1:
        d["fio_company"] = txt
        forms[uid]["step"] = 2
        await update.message.reply_text(T[lang]["ask_phone"], parse_mode="HTML"); return

    if st == 2:
        normalized, weird = normalize_phone(txt)
        d["phone"] = normalized
        if weird:
            await update.message.reply_text(T[lang]["bad_phone"], parse_mode="HTML")
        forms[uid]["step"] = 3
        await update.message.reply_text(T[lang]["ask_tg"], parse_mode="HTML"); return

    if st == 3:
        d["tg"] = txt.lstrip("@")
        forms[uid]["step"] = 4
        await update.message.reply_text(T[lang]["ask_task"], parse_mode="HTML"); return

    if st == 4:
        d["task"] = txt
        forms[uid]["step"] = 5
        await update.message.reply_text(T[lang]["ask_email"], parse_mode="HTML"); return

    if st == 5:
        d["email"] = txt
        if not email_valid(d["email"]):
            await update.message.reply_text(T[lang]["bad_email"], parse_mode="HTML")
        forms[uid]["step"] = 6
        kb = InlineKeyboardMarkup([[
            InlineKeyboardButton(T[lang]["btn_confirm"], callback_data="form:confirm"),
            InlineKeyboardButton(T[lang]["btn_cancel"], callback_data="form:cancel"),
        ]])
        await update.message.reply_text(
            f"{T[lang]['ask_confirm']}\n\n{render_card('', d)}",
            parse_mode="HTML", reply_markup=kb
        ); return

async def on_form_control(update: Update, context: ContextTypes.DEFAULT_TYPE):
    q = update.callback_query
    await q.answer()
    user = q.from_user
    maybe_auto_claim_admin(user)
    uid = user.id
    lang = get_lang(uid)
    d = forms.get(uid, {}).get("data", {})

    if q.data == "form:cancel":
        forms.pop(uid, None)
        await q.edit_message_text(T[lang]["cancelled"], parse_mode="HTML", reply_markup=main_menu(lang)); return

    if q.data == "form:confirm":
        # назначаем ID
        req_id = next_request_id()

        # 1) Запись в Excel
        try:
            excel_append_new(lang, req_id, d, user)
        except Exception as e:
            log.error("Excel append failed: %s", e)

        # 2) Отправка админу и (опционально) в командный чат
        text = f"{T[lang]['sent_admin']}\n\n{render_card(req_id, d)}"
        kb = InlineKeyboardMarkup([[
            InlineKeyboardButton("✅ В работе", callback_data=f"st:{req_id}:INPROG"),
            InlineKeyboardButton("⏳ Ожидание", callback_data=f"st:{req_id}:WAIT"),
            InlineKeyboardButton("❌ Отказ",   callback_data=f"st:{req_id}:REJ"),
        ]])

        if ADMIN_ID:
            try:
                await context.bot.send_message(chat_id=ADMIN_ID, text=text, parse_mode="HTML", reply_markup=kb)
            except Exception as e:
                log.error("Send to admin failed: %s", e)
        else:
            await q.message.reply_text(T[lang]["not_admin"])

        admin_chat = get_admin_chat_id()
        if admin_chat:
            try:
                await context.bot.send_message(chat_id=admin_chat, text=text, parse_mode="HTML", reply_markup=kb)
            except Exception as e:
                log.error("Send to team chat failed: %s", e)

        # 3) Ответ пользователю
        await q.edit_message_text(T[lang]["sent_user"], parse_mode="HTML", reply_markup=main_menu(lang))
        forms.pop(uid, None); return

# ── Запуск (webhook на Render, polling локально) ─────────────────────────────
def run(app):
    base_url = os.getenv("RENDER_EXTERNAL_URL")
    port = int(os.getenv("PORT", "10000"))
    if base_url:
        path = f"/webhook/{BOT_TOKEN}"
        webhook_url = f"{base_url}{path}"
        print(f">>> Using webhook on {webhook_url}")
        app.run_webhook(
            listen="0.0.0.0",
            port=port,
            url_path=path,
            webhook_url=webhook_url,
            allowed_updates=Update.ALL_TYPES,
        )
    else:
        print(">>> Using polling (local run)")
        app.run_polling(allowed_updates=Update.ALL_TYPES)

def main():
    bootstrap_admin_from_env()
    app = ApplicationBuilder().token(BOT_TOKEN).build()

    app.add_handler(CommandHandler("start", cmd_start))
    app.add_handler(CommandHandler("admin", cmd_admin))
    app.add_handler(CommandHandler("whoami", cmd_whoami))
    app.add_handler(CommandHandler("export", cmd_export))
    app.add_handler(CommandHandler("clear", cmd_clear))
    app.add_handler(CommandHandler("stats", cmd_stats))
    app.add_handler(CommandHandler("list", cmd_list))

    app.add_handler(CallbackQueryHandler(on_callback, pattern="^(set_lang:(ru|uz|en)|lang:open|form:start|st:.+)$"))
    app.add_handler(CallbackQueryHandler(on_form_control, pattern="^form:(confirm|cancel)$"))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, on_text))

    run(app)

if __name__ == "__main__":
    print(">>> Rubicon bot booting…")
    main()
